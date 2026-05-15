"""
Экспортёр Excel-отчётов.
Формирует профессиональный .xlsx отчёт с гиперссылками,
условным форматированием и автофильтром.
"""

from pathlib import Path
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from modules.utils.models import SearchTask, ReportRow
from modules.utils.helpers import availability_text
from config import OUTPUT_DIR, REPORT_COLUMNS, EXCEL_COLORS
from modules.logging.logger import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """Формирует Excel-отчёт по результатам поиска."""

    def export(self, tasks: list[SearchTask], region: str) -> Path:
        """
        Создаёт Excel-отчёт и возвращает путь к файлу.
        """
        rows = self._build_rows(tasks)
        file_path = self._make_file_path(region)
        self._write_excel(rows, file_path, region)
        logger.info("Отчёт сформирован: %s (%d строк)", file_path.name, len(rows))
        return file_path

    # ── Формирование строк отчёта ─────────────────────────────────────────────

    def _build_rows(self, tasks: list[SearchTask]) -> list[ReportRow]:
        """
        Для каждой позиции берём лучшее предложение (первое после ранжирования).
        Если предложений нет — строку всё равно включаем с пустыми полями.
        """
        rows: list[ReportRow] = []
        number = 0

        for task in tasks:
            spec = task.spec_item
            best = task.results[0] if task.results else None

            number += 1
            price = best.price if best else None
            qty = spec.quantity

            rows.append(ReportRow(
                number=number,
                name=spec.name,
                specs=spec.specs or "",
                unit=spec.unit or "",
                quantity=qty,
                price=price,
                total_price=(price * qty if price and qty else None),
                supplier_name=best.supplier_name if best else "",
                supplier_address=best.address if best else "",
                supplier_url=best.url if best else "",
                availability=availability_text(best.availability) if best else "Не найдено",
                note=task.error or (f"Найдено вариантов: {len(task.results)}" if task.results else "Не найдено"),
            ))

        return rows

    # ── Запись Excel ──────────────────────────────────────────────────────────

    def _write_excel(self, rows: list[ReportRow], path: Path, region: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ поставщиков"

        # Стили
        styles = self._build_styles()

        # Заголовок документа
        self._write_title(ws, region)

        # Заголовки колонок
        header_row = 3
        for col_idx, col_name in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["header_align"]
            cell.border = styles["thin_border"]

        # Данные
        for row_idx, row in enumerate(rows, start=header_row + 1):
            is_alt = (row_idx - header_row) % 2 == 0
            self._write_data_row(ws, row_idx, row, styles, is_alt)

        # Автофильтр
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(REPORT_COLUMNS))}{header_row + len(rows)}"

        # Закрепить заголовок
        ws.freeze_panes = f"A{header_row + 1}"

        # Ширина колонок
        self._set_column_widths(ws)

        # Лист с лучшими предложениями (топ-3 для каждой позиции)
        self._write_alternatives_sheet(wb, tasks=[], rows=rows)

        wb.save(str(path))

    def _write_title(self, ws, region: str) -> None:
        """Заголовок отчёта в первых двух строках."""
        ws.merge_cells(f"A1:{get_column_letter(len(REPORT_COLUMNS))}1")
        title_cell = ws["A1"]
        title_cell.value = f"Анализ поставщиков ТМЦ | Регион: {region}"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f"A2:{get_column_letter(len(REPORT_COLUMNS))}2")
        date_cell = ws["A2"]
        date_cell.value = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        date_cell.font = Font(italic=True, size=10, color="595959")
        date_cell.alignment = Alignment(horizontal="center")

    def _write_data_row(
        self, ws, row_idx: int, row: ReportRow, styles: dict, is_alt: bool
    ) -> None:
        """Записывает одну строку данных с форматированием."""
        fill = styles["alt_fill"] if is_alt else styles["white_fill"]
        border = styles["thin_border"]
        align_center = styles["data_align_center"]
        align_left = styles["data_align_left"]

        values = [
            row.number,
            row.name,
            row.specs,
            row.unit,
            row.quantity,
            row.price,
            row.total_price,
            row.supplier_name,
            row.supplier_address,
            row.supplier_url,  # URL — запишем как гиперссылку
            row.availability,
            row.note,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if col_idx != 10 else None)
            cell.fill = fill
            cell.border = border

            # Форматирование по типу колонки
            if col_idx in (1, 4, 5, 11):  # №, Ед.изм, Кол-во, Наличие
                cell.alignment = align_center
            elif col_idx in (6, 7):  # Цены
                cell.alignment = align_center
                if value:
                    cell.number_format = "#,##0.00"
            else:
                cell.alignment = align_left

            # Гиперссылка для URL
            if col_idx == 10 and row.supplier_url:
                cell.value = "Открыть →"
                cell.hyperlink = row.supplier_url
                cell.font = Font(color=EXCEL_COLORS["hyperlink_color"], underline="single")

        # Выделяем лучшие предложения (есть цена и в наличии)
        if row.price and row.availability == "В наличии":
            for col_idx in range(1, len(REPORT_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = styles["best_fill"]

    def _write_alternatives_sheet(self, wb, tasks: list, rows: list[ReportRow]) -> None:
        """Создаёт второй лист — краткая сводка."""
        ws2 = wb.create_sheet(title="Сводка")
        headers = ["№", "Наименование", "Лучшая цена (UZS)", "Поставщик", "Ссылка"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=EXCEL_COLORS["header_bg"])

        for row_idx, row in enumerate(rows, start=2):
            ws2.cell(row=row_idx, column=1, value=row.number)
            ws2.cell(row=row_idx, column=2, value=row.name)
            price_cell = ws2.cell(row=row_idx, column=3, value=row.price)
            price_cell.number_format = "#,##0.00"
            ws2.cell(row=row_idx, column=4, value=row.supplier_name)
            if row.supplier_url:
                link_cell = ws2.cell(row=row_idx, column=5, value="→")
                link_cell.hyperlink = row.supplier_url
                link_cell.font = Font(color=EXCEL_COLORS["hyperlink_color"], underline="single")

        for col in ws2.columns:
            ws2.column_dimensions[get_column_letter(col[0].column)].width = 25

    # ── Утилиты форматирования ────────────────────────────────────────────────

    @staticmethod
    def _build_styles() -> dict:
        thin = Side(style="thin", color="CCCCCC")
        return {
            "header_font": Font(bold=True, color=EXCEL_COLORS["header_fg"], size=10),
            "header_fill": PatternFill("solid", fgColor=EXCEL_COLORS["header_bg"]),
            "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "data_align_center": Alignment(horizontal="center", vertical="top", wrap_text=True),
            "data_align_left": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "thin_border": Border(left=thin, right=thin, top=thin, bottom=thin),
            "alt_fill": PatternFill("solid", fgColor=EXCEL_COLORS["alt_row_bg"]),
            "white_fill": PatternFill("solid", fgColor="FFFFFF"),
            "best_fill": PatternFill("solid", fgColor=EXCEL_COLORS["best_offer_bg"]),
        }

    @staticmethod
    def _set_column_widths(ws) -> None:
        """Устанавливает ширину колонок под контент."""
        widths = [5, 35, 25, 8, 10, 16, 18, 25, 25, 12, 12, 20]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Высота строки заголовка
        ws.row_dimensions[3].height = 40

    @staticmethod
    def _make_file_path(region: str) -> Path:
        """Генерирует имя файла отчёта."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_region = region.replace(" ", "_").replace("/", "-")
        return OUTPUT_DIR / f"report_{safe_region}_{timestamp}.xlsx"
